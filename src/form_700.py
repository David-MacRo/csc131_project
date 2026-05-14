from openpyxl import load_workbook
from config import DATA_PATH, SLASH, FORM_700_NAME

filename = f"{DATA_PATH}{SLASH}form_700{SLASH}{FORM_700_NAME}"


def load_names() -> list:
    #TODO load file in future
    form_names = ["Schedule A1", "Schedule A-2", "Schedule C - Income Section"]
    names = []
    form = load_workbook(filename)
    for sheet in form:
        if sheet.title in form_names:
            for row in sheet:
                if row[1].value is not None and row[1].value != "First Name":
                    name = f"{row[1].value} {row[0].value}"
                    if name not in names:
                        names.append(name)
    return names

def load_interests() -> list:
    #TODO load file in future
    form_names = ["Schedule A1", "Schedule A-2", "Schedule C - Income Section"]
    interests = []
    # interests = ["Landpaths"]
    form = load_workbook(filename)
    for sheet in form:
        if sheet.title in form_names:
            for row in sheet:
                if row[11].value is not None and row[11].value != "1. Income Received" and "business entity" not in row[11].value.lower() and "NAME OF SOURCE" not in row[11].value:
                    interest = row[11].value
                    if interest not in interests:
                        interests.append(interest)
    return interests

def load_name_interest_dict() -> dict:
    #TODO load file in future
    form_names = ["Schedule A1", "Schedule A-2", "Schedule C - Income Section"]
    form_700_data = {}
    form = load_workbook(filename)
    for sheet in form:
        if sheet.title in form_names:
            for row in sheet:
                if row[1].value is not None and row[1].value != "First Name":
                    name = f"{row[1].value} {row[0].value}"
                    interest = row[11].value
                    form_700_data.update({interest : name})
    print(form_700_data)

    return form_700_data

